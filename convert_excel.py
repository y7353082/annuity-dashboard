#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将Excel历史数据转换为JSON格式，导入到annuity-collector.html
"""

import json
import re
from openpyxl import load_workbook
from pathlib import Path

# 映射表
PLAN_MAP = {
    '单一计划': 'single',
    '集合计划': 'pooled',
    '其他计划': 'other',
}

TYPE_MAP = {
    '固定收益类': 'fixed',
    '含权益类': 'equity',
}

# 机构名称标准化（处理可能的差异）
INSTitution_MAP = {
    '太平养老': '太平养老',
    '华夏基金': '华夏基金',
    '银华基金': '银华基金',
    '招商基金': '招商基金',
    '中金公司': '中金公司',
    '平安养老': '平安养老',
    '泰康资产': '泰康资产',
    '国寿养老': '国寿养老',
    '人保养老': '人保养老',
    '南方基金': '南方基金',
    '易方达': '易方达',
    '博时基金': '博时基金',
    '中信证券': '中信证券',
    '华泰资产': '华泰资产',
    '国泰基金': '国泰基金',
    '工银瑞信': '工银瑞信',
    '长江养老': '长江养老',
    '富国基金': '富国基金',
    '嘉实基金': '嘉实基金',
    '海富通': '海富通',
    '建信养老': '建信养老',
    '新华养老': '新华养老',
}


def parse_quarter(q_text):
    """解析季度文本，如'2022年1季度' -> '2022Q1'"""
    match = re.match(r'(\d{4})年(\d)季度', q_text)
    if match:
        return f"{match.group(1)}Q{match.group(2)}"
    return q_text


def convert_excel_to_json(excel_path, output_dir):
    wb = load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    print(f"表头: {header}")
    print(f"数据行数: {len(data_rows)}")

    # 按季度分组
    quarters_data = {}

    for row in data_rows:
        q_text, inst_name, plan, type_, count, nav, ret = row

        quarter = parse_quarter(q_text)
        if quarter not in quarters_data:
            quarters_data[quarter] = {}

        # 标准化机构名称
        inst_key = inst_name
        for std_name in INSTitution_MAP.values():
            if std_name in inst_name:
                inst_key = std_name
                break

        if inst_key not in quarters_data[quarter]:
            quarters_data[quarter][inst_key] = {}

        plan_key = PLAN_MAP.get(plan, plan)
        type_key = TYPE_MAP.get(type_, type_)
        category_key = f"{plan_key}_{type_key}"

        def clean_num(val):
            if val is None:
                return "0"
            if isinstance(val, (int, float)):
                return str(float(val))
            # Handle string numbers (may contain Unicode minus sign)
            s = str(val).strip().replace('−', '-').replace('−', '-')
            try:
                return str(float(s))
            except ValueError:
                return "0"

        quarters_data[quarter][inst_key][category_key] = {
            "count": clean_num(count),
            "nav": clean_num(nav),
            "return": clean_num(ret),
        }

    # 保存为JSON文件
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    for quarter, data in sorted(quarters_data.items()):
        output_file = output_dir / f"annuity_data_{quarter}.json"
        export_obj = {
            "quarter": quarter,
            "exportDate": "2026-05-20",
            "source": "excel_import",
            "data": data
        }
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(export_obj, f, ensure_ascii=False, indent=2)
        print(f"已保存: {output_file} ({len(data)} 家机构)")

    return quarters_data


if __name__ == "__main__":
    excel_path = r"D:\工作\权益投资\BB排名统计-每季度\录入系统\2022-2025年信息披露数据.xlsx"
    output_dir = r"D:\工作\权益投资\BB排名统计-每季度\录入系统\json_export"

    quarters_data = convert_excel_to_json(excel_path, output_dir)

    # 打印汇总
    print("\n" + "=" * 60)
    print("转换完成")
    print("=" * 60)
    for quarter in sorted(quarters_data.keys()):
        inst_count = len(quarters_data[quarter])
        total_records = sum(len(v) for v in quarters_data[quarter].values())
        print(f"  {quarter}: {inst_count} 家机构, {total_records} 条记录")
    print("=" * 60)
    print(f"\nJSON文件已保存到: {output_dir}")
    print("在 annuity-collector.html 中点击「导入JSON」即可导入")
